import openpyxl

# 워크북 인스턴스 객체 => excel 파일 생성  (엑셀은 워크북이라고 불림)

# Workbook(워크북): 엑셀 "파일" 하나를 나타내는 객체. openpyxl.Workbook()으로 새 엑셀 파일을 메모리상에 생성.
wb = openpyxl.Workbook()
# 활성화된 워크북에 워크시트 객체 => 시트 만들기
sheet = wb.active
sheet.title = '회원정보'

# 헤더 컬럼
header_titles = ['아이디', '전화번호']
for idx, title in enumerate(header_titles):
    sheet.cell(row=1, column=idx+1, value=title)

# 내용 저장
members = [('happy', '010-1234-5678'), ('smile', '010-5555-5555')]

row_num = 2

for r, member in enumerate(members):
    for c, v in enumerate(member):
        sheet.cell(row=row_num, column=c+1, value=v)
    row_num+=1

wb.save('members.xlsx')
wb.close()